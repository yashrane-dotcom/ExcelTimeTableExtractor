"""
parser_engine.py — Core Excel parsing intelligence (v4.0)

Pipeline:
  1. Load workbook with openpyxl (preserves merged cells).
  2. Expand merged cells → full flat grid (no empty holes).
  3. Detect structure: day_col, time_col, class/division header cols.
  4. Discover teacher codes (whitelist + frequency heuristic).
  5. For each cell → extract_entities() → (teacher, division, room).
  6. Build structured schedules keyed by teacher / division / batch.
  7. Lab merging: consecutive same-subject lab slots → [2hr] / [3hr].
  8. Compact output: {"MON": {"8:30": ["UPM SE-A 301", "DKP TE-B 204"]}}
  9. Full JSON output: {"teacher": {...}, "division": {...}}
"""

from __future__ import annotations

import re
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger(__name__)

# ─── Faculty map ──────────────────────────────────────────────────────────────

FACULTY_MAP: dict[str, str] = {
    "UPM": "Dr. Umesh Moharil",      "MRY": "Dr. Meghna Yashwante",
    "MDB": "Ms. Manisha Bhise",      "AGP": "Dr. Amita Pal",
    "AGD": "Dr. Anil Darekar",       "BDP": "Dr. B D Patil",
    "PSD": "Dr. Pratibha Desai",     "PMN": "Dr. Poonam Nakhate",
    "MS":  "Mr. Mukesh Sharma",      "RBM": "Mr. Rahul Mali",
    "HDV": "Mr. Harshal Vaidya",     "PG":  "Mr. Pankaj Gaur",
    "VVK": "Mr. Vishal Kulkarni",    "RPD": "Mr. R P Dharmale",
    "SIB": "Mr. Sanket Barde",       "SG":  "Dr. Sandhya Gadge",
    "NG":  "Mr. Nikhil Gurav",       "PVM": "Mrs. Pallavi Munde",
    "SK":  "Ms. Sheetal Khande",     "CJ":  "Dr. Chhaya Joshi",
    "TP":  "Mr. Tukaram Patil",      "ST":  "Ms. Shilpa Tambe",
    "NV":  "Mrs. Neha Verma",        "SM":  "Mrs. Sonali Murumkar",
    "SB":  "Ms. Swati Bagade",       "SD":  "Mr. Shankar Deshmukh",
    "MPP": "Mr. Martand Pandagale",  "DKP": "Mr. Deepak Patil",
    "ABC": "Staff",
}

SKIP_TOKENS: frozenset[str] = frozenset({
    "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN",
    "THE", "AND", "FOR", "NOT", "ARE", "WAS", "HAS",
    "LAB", "LEC", "TH", "DAY", "TIME", "AM", "PM",
    "NO", "ID", "PR", "TD", "LE", "ROOM", "CLASS",
    "DIVISION", "SLOT", "BREAK", "LUNCH", "FREE", "PRACTICAL",
    "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY",
    "A", "B", "C", "D", "E", "F", "G", "H",
})

DAYS_SHORT     = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
DAYS_FULL_LIST = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
FIXED_DAYS     = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]

SLOT_CONFIG = [
    {"key": "8:30",  "label": "8:30",  "type": "slot"},
    {"key": "9:30",  "label": "9:30",  "type": "slot"},
    {"key": "BRK",   "label": "Break", "type": "break"},
    {"key": "10:45", "label": "10:45", "type": "slot"},
    {"key": "11:45", "label": "11:45", "type": "slot"},
    {"key": "12:45", "label": "12:45", "type": "slot"},
    {"key": "LCH",   "label": "Lunch", "type": "lunch"},
    {"key": "1:30",  "label": "1:30",  "type": "slot"},
    {"key": "2:30",  "label": "2:30",  "type": "slot"},
    {"key": "3:30",  "label": "3:30",  "type": "slot"},
]
LECTURE_SLOTS = [s["key"] for s in SLOT_CONFIG if s["type"] == "slot"]

DIVISION_CONFIG: dict[str, dict] = {
    "A": {"name": "FE-A (Comp)",  "batches": ["A1", "A2", "A3"]},
    "B": {"name": "FE-B (Civil)", "batches": ["B1", "B2", "B3"]},
    "C": {"name": "FE-C (Comp)",  "batches": ["C1", "C2", "C3"]},
    "D": {"name": "FE-D (Mech)",  "batches": ["D1", "D2", "D3"]},
    "E": {"name": "FE-E (AI&DS)", "batches": ["E1", "E2", "E3"]},
    "F": {"name": "FE-F (Ro&AI)", "batches": ["F1", "F2", "F3"]},
    "G": {"name": "FE-G (AI&DS)", "batches": ["G1", "G2", "G3"]},
    "H": {"name": "FE-H (MTRX)",  "batches": ["H1", "H2", "H3"]},
}

# ─── Regex Patterns ───────────────────────────────────────────────────────────

TEACHER_CODE_RE = re.compile(r'\b([A-Z]{2,5})\b')

DIVISION_RE = re.compile(
    r'\b(?:'
    r'(?:FE|SE|TE|BE)[\s\-]?([A-H](?:\d+)?)'
    r'|'
    r'([A-H])[\s\-]?(?:DIV(?:ISION)?|SEC(?:TION)?)?'
    r')\b',
    re.IGNORECASE
)

DIV_WITH_PREFIX_RE = re.compile(
    r'\b((?:FE|SE|TE|BE)[\s\-]?[A-H](?:\d+)?)\b',
    re.IGNORECASE
)

ROOM_RE = re.compile(
    r'(?:'
    r'(?:room|rm|lab|hall|cr|sr|lr|gr|cb|blk|block)\s*[\-\.]?\s*(\d{1,4}[A-Z]?)'
    r'|'
    r'\b(\d{3,4}[A-Z]?)\b'
    r')',
    re.IGNORECASE
)

LAB_ROOM_RE = re.compile(
    r'\b(?:lab|laboratory)[\s\-]?(\d+[A-Z]?)\b'
    r'|'
    r'\b([A-Z]\d)[\s\-]?(?:lab|laboratory)[\s\-]?(\d+)?\b',
    re.IGNORECASE
)

BATCH_RE = re.compile(r'\b([A-H]\d)\b', re.IGNORECASE)

TIME_RE = re.compile(r'(\d{1,2})[:\.](\d{2})(\s*(AM|PM))?', re.IGNORECASE)

LAB_KEYWORD_RE = re.compile(r'\b(lab|pr|practical|laboratory)\b', re.IGNORECASE)


# ─── Entity Extraction Core ───────────────────────────────────────────────────

def extract_entities(cell_text: str, known_teachers: set[str] | None = None) -> list[dict]:
    """
    PRIMARY EXTRACTION FUNCTION.

    From a single cell string, extract all (teacher, division, room) triplets.
    Returns a list of dicts:
      [{"teacher": "UPM", "division": "SE-A", "room": "301",
        "is_lab": False, "formatted": "UPM SE-A 301"}, ...]

    Handles:
      - Multiple teachers/classes in one cell
      - Lab rooms (Lab-204)
      - Messy real-world formats
      - Cells with no room or no division
    """
    if not cell_text or not cell_text.strip():
        return []

    known_teachers = known_teachers or set(FACULTY_MAP.keys())
    segments = _split_cell_segments(cell_text)
    results = []

    for seg in segments:
        seg = seg.strip()
        if not seg or len(seg) < 3:
            continue
        results.extend(_extract_from_segment(seg, known_teachers))

    # Deduplicate by formatted string
    seen: set[str] = set()
    deduped = []
    for e in results:
        key = e["formatted"]
        if key not in seen:
            seen.add(key)
            deduped.append(e)

    return deduped


def _extract_from_segment(seg: str, known_teachers: set[str]) -> list[dict]:
    """Extract all (teacher, division, room) entities from one segment."""
    teachers  = _extract_teachers(seg, known_teachers)
    divisions = _extract_divisions(seg)
    rooms     = _extract_rooms(seg)
    is_lab    = bool(LAB_KEYWORD_RE.search(seg))

    results: list[dict] = []

    if not teachers:
        return results

    if not divisions:
        for t in teachers:
            results.append(_make_entity(t, "", rooms[0] if rooms else "", is_lab))
        return results

    if len(teachers) == len(divisions):
        for t, d in zip(teachers, divisions):
            results.append(_make_entity(t, d, rooms[0] if rooms else "", is_lab))
    elif len(teachers) == 1:
        for d in divisions:
            results.append(_make_entity(teachers[0], d, rooms[0] if rooms else "", is_lab))
    else:
        for i, t in enumerate(teachers):
            d = divisions[i] if i < len(divisions) else divisions[-1]
            r = rooms[i] if i < len(rooms) else (rooms[0] if rooms else "")
            results.append(_make_entity(t, d, r, is_lab))

    return results


def _make_entity(teacher: str, division: str, room: str, is_lab: bool) -> dict:
    """Build a clean entity dict with a formatted string."""
    parts = [p for p in [teacher, division, room] if p]
    return {
        "teacher":   teacher,
        "division":  division,
        "room":      room,
        "is_lab":    is_lab,
        "formatted": " ".join(parts),
    }


def _extract_teachers(text: str, known_teachers: set[str]) -> list[str]:
    """Extract teacher codes from text (whitelist-strict, paren-first)."""
    found: list[str] = []
    seen:  set[str]  = set()
    text_up = text.upper()

    # Parenthesised codes are strongest signal
    for code in re.findall(r'\(([A-Z]{2,5})\)', text_up):
        if code in known_teachers and code not in seen:
            seen.add(code)
            found.append(code)

    for m in TEACHER_CODE_RE.finditer(text_up):
        code = m.group(1)
        if code in SKIP_TOKENS:
            continue
        if code in known_teachers and code not in seen:
            seen.add(code)
            found.append(code)

    return found


def _looks_like_teacher_code(code: str, context: str) -> bool:
    """Legacy helper — kept for compatibility."""
    if len(code) < 2 or len(code) > 5:
        return False
    if code in SKIP_TOKENS or BATCH_RE.match(code) or not code.isalpha():
        return False
    if code in {"FE", "SE", "TE", "BE"}:
        return False
    return True


def _extract_divisions(text: str) -> list[str]:
    """Extract division strings like SE-A, TE-B, FE-C, returning normalised forms."""
    found: list[str] = []
    seen:  set[str]  = set()

    for m in DIV_WITH_PREFIX_RE.finditer(text):
        norm = _normalise_division(m.group(1))
        if norm and norm not in seen:
            seen.add(norm)
            found.append(norm)

    if not found:
        for m in DIVISION_RE.finditer(text):
            letter = ((m.group(1) or m.group(2)) or "").upper()
            if letter and letter not in seen and letter in "ABCDEFGH":
                seen.add(letter)
                found.append(letter)

    return found


def _normalise_division(raw: str) -> str:
    r = raw.strip().upper()
    r = re.sub(r'\s*-\s*', '-', r)
    r = re.sub(r'\s+', '-', r)
    m = re.match(r'^(FE|SE|TE|BE)([A-H]\d*)$', r)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return r


def _extract_rooms(text: str) -> list[str]:
    """Extract room numbers, prioritising lab-specific patterns."""
    found: list[str] = []
    seen:  set[str]  = set()

    for m in LAB_ROOM_RE.finditer(text):
        num = m.group(1) or m.group(3)
        if num:
            room = f"Lab-{num}"
            if room not in seen:
                seen.add(room)
                found.append(room)

    for m in ROOM_RE.finditer(text):
        named = m.group(1)
        bare  = m.group(2)
        room  = (named or bare or "").strip()
        if not room:
            continue
        if f"Lab-{room}" in seen:
            continue
        if len(room) == 4 and room[:2] in ("19", "20"):
            continue
        if room not in seen:
            seen.add(room)
            found.append(room)

    return found


# ─── Cell splitting ───────────────────────────────────────────────────────────

def _split_cell_segments(raw: str) -> list[str]:
    """
    Split one cell's text into individual schedule segments.
    Handles newline-delimited entries and comma-splits outside parentheses.
    """
    lines = [l.strip() for l in re.split(r'[\n\r]+', raw) if l.strip()]
    result = []
    for line in lines:
        result.extend(_split_comma_outside_parens(line))
    return [s.strip() for s in result if s.strip()] or [raw.strip()]


def _split_comma_outside_parens(s: str) -> list[str]:
    parts, depth, start = [], 0, 0
    for i, ch in enumerate(s):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 0:
            parts.append(s[start:i])
            start = i + 1
    parts.append(s[start:])
    return parts


# ─── Utility helpers ──────────────────────────────────────────────────────────

def normalise_time(raw: str) -> str:
    raw = raw.strip()
    m = TIME_RE.search(raw)
    if not m:
        return ""
    h, mn = int(m.group(1)), m.group(2)
    ampm = (m.group(4) or "").upper()
    if ampm == "PM" and h < 12:
        h += 12
    if ampm == "AM" and h == 12:
        h = 0
    return f"{h}:{mn}"


def time_to_mins(t: str) -> int:
    m = re.match(r'(\d{1,2}):(\d{2})', t)
    if not m:
        return 9999
    h, mn = int(m.group(1)), int(m.group(2))
    if h < 7:
        h += 12
    return h * 60 + mn


def match_day(val: str) -> str | None:
    v = val.strip().upper()
    for d in DAYS_SHORT:
        if v == d:
            return d
    for i, d in enumerate(DAYS_FULL_LIST):
        if v == d or v.startswith(d):
            return DAYS_SHORT[i]
    return None


def resolve_slot_key(time_norm: str) -> str | None:
    mins = time_to_mins(time_norm)
    best, best_diff = None, 99999
    for key in LECTURE_SLOTS:
        diff = abs(time_to_mins(key) - mins)
        if diff < best_diff:
            best_diff, best = diff, key
    return best if best_diff <= 20 else None


def contains_teacher(text: str, code: str) -> bool:
    pattern = re.compile(
        rf'(?:^|[^A-Za-z]){re.escape(code)}(?:[^A-Za-z]|$)', re.IGNORECASE
    )
    return bool(pattern.search(text))


def clean_subject(raw_seg: str, remove_codes: list[str]) -> str:
    """Strip teacher codes, batch codes, room numbers, and junk — leave only subject."""
    s = raw_seg
    s = BATCH_RE.sub('', s)
    s = re.sub(r'\(\s*[A-H]\d\s*\)', '', s, flags=re.IGNORECASE)
    for code in remove_codes:
        if not code:
            continue
        pattern = re.compile(
            rf'(?:^|[^A-Za-z0-9]){re.escape(code)}(?:[^A-Za-z0-9]|$)',
            re.IGNORECASE
        )
        s = pattern.sub(' ', s)
        s = re.sub(rf'\(\s*{re.escape(code)}\s*\)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b(Lab|Pr|Lec|Practical|TD|LE|Room|Rm)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{3,}\b', '', s)
    s = re.sub(r'\(([^)]*)\)', _clean_paren, s)
    s = re.sub(r'^[\s\-:\,–—]+', '', s)
    s = re.sub(r'[\s\-:\,–—]+$', '', s)
    skip_pattern = r'\b(?:' + '|'.join(re.escape(t) for t in sorted(SKIP_TOKENS, key=len, reverse=True)) + r')\b'
    s = re.sub(skip_pattern, ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\s{2,}', ' ', s).strip()
    return s


def _clean_paren(m: re.Match) -> str:
    inner = m.group(1).strip()
    toks = re.split(r'[\s,·]+', inner)
    if all(re.match(r'^[A-Z]{2,5}$', t) or re.match(r'^[A-H]\d$', t) or not t
           for t in toks):
        return ''
    return m.group(0)


def extract_teacher_codes(raw: str) -> list[str]:
    return [m.group(1) for m in TEACHER_CODE_RE.finditer(raw.upper())
            if m.group(1) not in SKIP_TOKENS]


# ─── Lab merge helpers ────────────────────────────────────────────────────────

def _compact_lab_label(hours: int) -> str:
    return f"[{hours}hr]" if hours > 1 else ""


def _merge_compact_labs(
    day_slots: dict[str, list[str]]
) -> dict[str, list[str]]:
    """
    Given a dict of {slot_key: [compact_strings]}, merge consecutive lab slots
    that share the same teacher+division into a single entry with [Xhr] suffix.

    E.g.
      "10:45": ["UPM SE-A Lab-204"]
      "11:45": ["UPM SE-A Lab-204"]
    becomes
      "10:45": ["UPM SE-A Lab-204 [2hr]"]
      "11:45": [REMOVED]
    """
    ordered = [k for k in LECTURE_SLOTS if k in day_slots]
    result:   dict[str, list[str]] = {k: list(v) for k, v in day_slots.items()}
    skip:     set[str]             = set()

    for i, key in enumerate(ordered):
        if key in skip:
            continue
        entries = result.get(key, [])
        if not entries:
            continue

        for entry in list(entries):
            # Only merge entries that look like labs (contain Lab- or match lab room heuristic)
            if not _entry_is_lab(entry):
                continue

            span   = 1
            cursor = i + 1
            while cursor < len(ordered):
                next_key     = ordered[cursor]
                next_entries = result.get(next_key, [])
                # Check if the identical compact string appears in next slot
                if entry in next_entries:
                    span   += 1
                    skip.add(next_key)
                    cursor  += 1
                else:
                    break

            if span > 1:
                # Replace entry in current slot with version that has [Xhr] suffix
                labelled = _attach_lab_label(entry, span)
                idx = result[key].index(entry)
                result[key][idx] = labelled
                # Remove merged slots entirely
                for s_key in ordered[i + 1: i + span]:
                    skip.add(s_key)

    # Drop skipped slots
    return {k: v for k, v in result.items() if k not in skip and v}


def _entry_is_lab(compact: str) -> bool:
    """Heuristic: entry is a lab if it contains 'Lab-' or the room starts with Lab."""
    return bool(re.search(r'\bLab[-\d]', compact, re.IGNORECASE))


def _attach_lab_label(compact: str, hours: int) -> str:
    """Append [Xhr] to a compact string if not already present."""
    if re.search(r'\[\d+hr\]', compact):
        return compact
    label = _compact_lab_label(hours)
    return f"{compact} {label}".strip() if label else compact


# ─── Full master timetable split ──────────────────────────────────────────────

def build_full_timetable_json(
    compact_schedule: dict[str, dict[str, list[str]]]
) -> dict:
    """
    Convert the compact master schedule into:
      {
        "teacher": {
          "UPM": {"MON": {"8:30": "UPM SE-A 301", "10:45": "UPM SE-A Lab-204 [2hr]"}}
        },
        "division": {
          "SE-A": {"MON": {"8:30": "UPM SE-A 301"}}
        }
      }

    Input is the output of TimetableParser.extract_compact_schedule() after
    lab merging has been applied.
    """
    teacher_tt:  dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))
    division_tt: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

    for day, slots in compact_schedule.items():
        for slot_key, entries in slots.items():
            for entry in entries:
                parts = entry.split()
                if not parts:
                    continue
                teacher  = parts[0] if len(parts) >= 1 else ""
                division = parts[1] if len(parts) >= 2 else ""
                # room + optional label are the remaining parts
                # (entry is already the full compact string)

                if teacher:
                    teacher_tt[teacher][day][slot_key] = entry
                if division:
                    division_tt[division][day][slot_key] = entry

    return {
        "teacher":  {t: dict(d) for t, d in teacher_tt.items()},
        "division": {d: dict(s) for d, s in division_tt.items()},
    }


# ─── Main Parser ──────────────────────────────────────────────────────────────

class TimetableParser:
    """
    Stateful parser: load once, query many times.
    v4: adds lab merging to compact schedule, full JSON split output,
        and improved merged-cell + structure detection.
    """

    FIXED_DAYS = FIXED_DAYS

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.grid: list[list[str]] = []
        self.row_count = 0
        self.col_count = 0
        self.sheet_names: list[str] = []
        self.day_col  = 0
        self.time_col = 1
        self.class_col_map: dict[int, str] = {}
        self.day_row_map: dict[str, list[tuple[int, str]]] = {}
        self.div_col_map: dict[str, set[int]] = {d: set() for d in DIVISION_CONFIG}
        self._teacher_occ: dict[str, int] = {}
        self._known_teachers: set[str] = set()
        self._merged_info: dict[tuple[int, int], dict[str, int]] = {}

    # ── Step 1: Load & expand merged cells ────────────────────────────────────

    def _load_grid(self) -> list[list[str]]:
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        self.sheet_names = wb.sheetnames
        ws = wb.active

        merge_map: dict[tuple[int, int], str] = {}
        self._merged_info = {}

        for merge_range in ws.merged_cells.ranges:
            tl_cell  = ws.cell(merge_range.min_row, merge_range.min_col)
            val      = str(tl_cell.value or "").strip()
            span_rows = merge_range.max_row - merge_range.min_row + 1
            span_cols = merge_range.max_col - merge_range.min_col + 1

            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merge_map[(row, col)] = val
                    is_owner = (row == merge_range.min_row and col == merge_range.min_col)
                    self._merged_info[(row - 1, col - 1)] = {
                        "owner_row": merge_range.min_row - 1,
                        "owner_col": merge_range.min_col - 1,
                        "span_rows": span_rows if is_owner else 0,
                        "span_cols": span_cols if is_owner else 0,
                    }

        grid: list[list[str]] = []
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            row_data: list[str] = []
            for c_idx, cell in enumerate(row, start=1):
                key = (r_idx, c_idx)
                row_data.append(merge_map[key] if key in merge_map
                                 else str(cell.value or "").strip())
            grid.append(row_data)
        return grid

    def _is_merged_continuation(self, row: int, col: int) -> bool:
        info = self._merged_info.get((row, col))
        return bool(info and (info["owner_row"], info["owner_col"]) != (row, col))

    def _merged_row_span(self, row: int, col: int) -> int:
        info = self._merged_info.get((row, col))
        if info and (info["owner_row"], info["owner_col"]) == (row, col):
            return max(1, info.get("span_rows", 1))
        return 1

    # ── Step 2: Detect column roles ───────────────────────────────────────────

    def _detect_columns(self):
        day_votes:  dict[int, int] = defaultdict(int)
        time_votes: dict[int, int] = defaultdict(int)

        for row in self.grid:
            for c, val in enumerate(row):
                v = val.strip().upper()
                if v in DAYS_SHORT or any(v == d or v.startswith(d) for d in DAYS_FULL_LIST):
                    day_votes[c] += 1
                if TIME_RE.match(v):
                    time_votes[c] += 1

        self.day_col  = max(day_votes,  key=day_votes.get,  default=0)
        self.time_col = max(time_votes, key=time_votes.get, default=1)
        log.info(f"[columns] day_col={self.day_col}  time_col={self.time_col}")

    # ── Step 3: Detect class/division columns ─────────────────────────────────

    def _detect_class_cols(self):
        header_rows_to_check = min(6, len(self.grid))
        for r in range(header_rows_to_check):
            row = self.grid[r]
            row_val = row[self.day_col].strip().upper() if self.day_col < len(row) else ""
            if row_val in {"", "CLASS", "DIVISION", "DAY"}:
                for c, val in enumerate(row):
                    if c in (self.day_col, self.time_col):
                        continue
                    v = val.strip()
                    if v and not v.isdigit():
                        self.class_col_map[c] = v
                        vu = v.upper()
                        for div in DIVISION_CONFIG:
                            if (vu == div or vu == f"FE-{div}" or vu == f"FE {div}"
                                    or vu.startswith(f"FE-{div} ")
                                    or f"({div})" in vu or vu.startswith(f"{div} (")):
                                self.div_col_map[div].add(c)

        if not self.class_col_map and self.grid:
            for c, val in enumerate(self.grid[0]):
                if c not in (self.day_col, self.time_col):
                    v = val.strip()
                    if v and not v.isdigit():
                        self.class_col_map[c] = v

        log.info(f"[class_cols] {self.class_col_map}")

    # ── Step 4: Map days to rows ───────────────────────────────────────────────

    def _build_day_row_map(self):
        cur_day = None
        self.day_row_map = {d: [] for d in FIXED_DAYS}
        for r, row in enumerate(self.grid):
            day_val  = row[self.day_col].strip()  if self.day_col  < len(row) else ""
            time_raw = row[self.time_col].strip()  if self.time_col < len(row) else ""
            dk = match_day(day_val.upper())
            if dk:
                cur_day = dk
            if cur_day and cur_day in FIXED_DAYS:
                t_norm = normalise_time(time_raw)
                if t_norm:
                    slot_key = resolve_slot_key(t_norm)
                    if slot_key:
                        self.day_row_map[cur_day].append((r, slot_key))
        log.info(f"[day_rows]  { {d: len(v) for d, v in self.day_row_map.items()} }")

    # ── Step 5: Discover teachers ─────────────────────────────────────────────

    def _discover_teachers(self):
        known = set(FACULTY_MAP.keys())
        occ: dict[str, int] = defaultdict(int)
        for row in self.grid:
            for cell in row:
                if not cell:
                    continue
                for code in extract_teacher_codes(cell):
                    if code in known:
                        self._known_teachers.add(code)
                    elif code not in SKIP_TOKENS and len(code) >= 2:
                        occ[code] += 1
        for code, count in occ.items():
            if count >= 2:
                self._known_teachers.add(code)
        self._teacher_occ = dict(occ)
        log.info(f"[teachers] found {len(self._known_teachers)}: {sorted(self._known_teachers)}")

    # ── Public orchestrator ────────────────────────────────────────────────────

    def parse(self):
        log.info(f"[parse] loading {self.filepath}")
        self.grid = self._load_grid()
        if not self.grid:
            raise ValueError("Workbook appears to be empty.")
        self.row_count = len(self.grid)
        self.col_count = max(len(r) for r in self.grid)
        self._detect_columns()
        self._detect_class_cols()
        self._build_day_row_map()
        self._discover_teachers()
        log.info("[parse] complete")

    # ── Public accessors ──────────────────────────────────────────────────────

    def get_teacher_codes(self) -> set[str]:
        return self._known_teachers

    def get_teachers_with_names(self) -> list[dict]:
        return [
            {"code": c, "display_name": self.faculty_display_name(c)}
            for c in sorted(self._known_teachers)
        ]

    def faculty_display_name(self, code: str) -> str:
        return FACULTY_MAP.get(code.upper(), code.upper())

    def get_divisions(self) -> list[str]:
        return sorted(DIVISION_CONFIG.keys())

    def slot_config_json(self) -> list[dict]:
        return SLOT_CONFIG

    # ── Compact schedule builder ──────────────────────────────────────────────

    def extract_compact_schedule(
        self, merge_labs: bool = True
    ) -> dict[str, dict[str, list[str]]]:
        """
        Build the full timetable as compact strings with lab merging.

        Returns:
          {
            "MON": {
              "8:30":  ["UPM SE-A 301"],
              "10:45": ["UPM SE-A Lab-204 [2hr]"],   # merged from 10:45+11:45
            },
            ...
          }
        """
        raw: dict[str, dict[str, list[str]]] = {
            day: {slot: [] for slot in LECTURE_SLOTS}
            for day in FIXED_DAYS
        }

        for day, row_entries in self.day_row_map.items():
            for row_idx, slot_key in row_entries:
                row = self.grid[row_idx]
                for c, cell_val in enumerate(row):
                    if c in (self.day_col, self.time_col) or not cell_val:
                        continue
                    if self._is_merged_continuation(row_idx, c):
                        continue
                    entities = extract_entities(cell_val, self._known_teachers)
                    for e in entities:
                        fmt = e["formatted"]
                        if fmt and fmt not in raw[day][slot_key]:
                            raw[day][slot_key].append(fmt)

        # Remove empty slots then optionally merge labs
        cleaned: dict[str, dict[str, list[str]]] = {}
        for day in FIXED_DAYS:
            non_empty = {s: v for s, v in raw[day].items() if v}
            if merge_labs:
                non_empty = _merge_compact_labs(non_empty)
            if non_empty:
                cleaned[day] = non_empty

        return cleaned

    # ── Full split JSON (teacher + division) ──────────────────────────────────

    def extract_full_timetable_json(self) -> dict:
        """
        Convert master timetable into per-teacher and per-division schedules.

        Returns:
          {
            "teacher": {
              "UPM": {
                "MON": {
                  "8:30":  "UPM SE-A 301",
                  "10:45": "UPM SE-A Lab-204 [2hr]"
                }
              }
            },
            "division": {
              "SE-A": {
                "MON": {"8:30": "UPM SE-A 301"}
              }
            }
          }
        """
        compact = self.extract_compact_schedule(merge_labs=True)
        return build_full_timetable_json(compact)

    # ── Teacher schedule (rich SlotEntry format) ──────────────────────────────

    def build_teacher_schedule(self, teacher_code: str) -> dict:
        schedule = self._empty_schedule()
        code_up  = teacher_code.upper()

        for day, row_entries in self.day_row_map.items():
            for row_idx, slot_key in row_entries:
                row = self.grid[row_idx]
                for c, cell_val in enumerate(row):
                    if c in (self.day_col, self.time_col) or not cell_val:
                        continue
                    if self._is_merged_continuation(row_idx, c):
                        continue
                    if not contains_teacher(cell_val, code_up):
                        continue
                    span = self._merged_row_span(row_idx, c)
                    for seg in _split_cell_segments(cell_val):
                        if not contains_teacher(seg, code_up):
                            continue
                        entry = self._parse_segment_for_teacher(seg, code_up)
                        if not entry:
                            continue
                        entry["class_div"] = self.class_col_map.get(c, "")
                        if span > 1:
                            entry["span"] = span
                        if not any(e["subject"] == entry["subject"] and e["type"] == entry["type"]
                                   for e in schedule[day][slot_key]):
                            schedule[day][slot_key].append(entry)

        merge_info = self._merge_labs(schedule)
        return self._schedule_to_response(schedule, merge_info)

    def _parse_segment_for_teacher(self, seg: str, teacher_code: str) -> dict | None:
        is_lab      = bool(LAB_KEYWORD_RE.search(seg))
        all_codes   = [m.group(1) for m in TEACHER_CODE_RE.finditer(seg.upper())
                       if m.group(1) not in SKIP_TOKENS]
        faculty_codes = [c for c in all_codes
                         if c != teacher_code
                         and not BATCH_RE.match(c)
                         and c not in SKIP_TOKENS]
        subject = clean_subject(seg, [teacher_code] + faculty_codes)
        if not subject or len(subject) < 2 or subject.upper() in SKIP_TOKENS:
            return None

        divisions = _extract_divisions(seg)
        rooms     = _extract_rooms(seg)

        return {
            "subject":       subject,
            "type":          "lab" if is_lab else "lec",
            "faculty_codes": faculty_codes,
            "class_div":     divisions[0] if divisions else "",
            "room":          rooms[0] if rooms else "",
            "compact":       f"{teacher_code} {divisions[0] if divisions else ''} {rooms[0] if rooms else ''}".strip(),
            "raw":           seg,
        }

    # ── Division schedule ─────────────────────────────────────────────────────

    def build_division_schedule(self, div: str) -> dict:
        schedule = self._empty_schedule()
        div_cols = self.div_col_map.get(div, set())

        for day, row_entries in self.day_row_map.items():
            for row_idx, slot_key in row_entries:
                row = self.grid[row_idx]
                for c, cell_val in enumerate(row):
                    if c in (self.day_col, self.time_col) or not cell_val:
                        continue
                    if self._is_merged_continuation(row_idx, c):
                        continue
                    if div_cols and c not in div_cols:
                        continue
                    col_label = self.class_col_map.get(c, "")
                    if not div_cols:
                        vu = col_label.upper()
                        if not (vu == div or f"FE-{div}" in vu or f"FE {div}" in vu):
                            continue
                    for seg in _split_cell_segments(cell_val):
                        entry = self._parse_generic_segment(seg)
                        if not entry:
                            continue
                        entry["class_div"] = col_label
                        span = self._merged_row_span(row_idx, c)
                        if span > 1:
                            entry["span"] = span
                        if not any(e["subject"] == entry["subject"] and e["type"] == entry["type"]
                                   for e in schedule[day][slot_key]):
                            schedule[day][slot_key].append(entry)

        merge_info = self._merge_labs(schedule)
        return self._schedule_to_response(schedule, merge_info)

    # ── Batch schedule ────────────────────────────────────────────────────────

    def build_batch_schedule(self, div: str, batch: str) -> dict:
        schedule  = self._empty_schedule()
        div_cols  = self.div_col_map.get(div, set())
        batch_up  = batch.upper()

        for day, row_entries in self.day_row_map.items():
            for row_idx, slot_key in row_entries:
                row = self.grid[row_idx]
                for c, cell_val in enumerate(row):
                    if c in (self.day_col, self.time_col) or not cell_val:
                        continue
                    if self._is_merged_continuation(row_idx, c):
                        continue
                    if not self._cell_belongs_to_batch(cell_val, div, batch_up, c, div_cols):
                        continue
                    for seg in _split_cell_segments(cell_val):
                        entry = self._parse_segment_for_batch(seg, div, batch_up)
                        if not entry:
                            continue
                        entry["class_div"] = self.class_col_map.get(c, "")
                        span = self._merged_row_span(row_idx, c)
                        if span > 1:
                            entry["span"] = span
                        if not any(e["subject"] == entry["subject"] and e["type"] == entry["type"]
                                   for e in schedule[day][slot_key]):
                            schedule[day][slot_key].append(entry)

        merge_info = self._merge_labs(schedule)
        return self._schedule_to_response(schedule, merge_info)

    def _cell_belongs_to_batch(self, cell: str, div: str, batch: str, col: int, div_cols: set) -> bool:
        if re.search(rf'(?:^|[^A-Za-z0-9]){re.escape(batch)}(?:[^A-Za-z0-9]|$)', cell, re.IGNORECASE):
            return True
        if not BATCH_RE.search(cell):
            if div_cols and col in div_cols:
                return True
            col_label = self.class_col_map.get(col, "").upper()
            if col_label and (col_label == div
                              or col_label.startswith(f"FE-{div}")
                              or col_label.startswith(f"FE {div}")):
                return True
        return False

    def _parse_segment_for_batch(self, seg: str, div: str, batch: str) -> dict | None:
        s = seg.strip()
        if not s or len(s) < 2:
            return None
        is_lab = bool(LAB_KEYWORD_RE.search(s))
        all_caps = [m.group(1) for m in TEACHER_CODE_RE.finditer(s.upper())]
        batch_codes_in_seg = [c for c in all_caps if BATCH_RE.match(c)]
        if batch_codes_in_seg and batch.upper() not in [c.upper() for c in batch_codes_in_seg]:
            return None
        faculty_codes = [c for c in all_caps
                         if not BATCH_RE.match(c)
                         and c not in SKIP_TOKENS
                         and (c in FACULTY_MAP or c in self._known_teachers)]
        subject = clean_subject(s, faculty_codes + batch_codes_in_seg)
        if not subject or len(subject) < 2 or subject.upper() in SKIP_TOKENS:
            return None
        rooms = _extract_rooms(s)
        return {
            "subject":       subject,
            "type":          "lab" if is_lab else "lec",
            "faculty_codes": faculty_codes,
            "class_div":     "",
            "room":          rooms[0] if rooms else "",
            "raw":           s,
        }

    def _parse_generic_segment(self, seg: str) -> dict | None:
        s = seg.strip()
        if not s or len(s) < 2:
            return None
        is_lab   = bool(LAB_KEYWORD_RE.search(s))
        all_caps = [m.group(1) for m in TEACHER_CODE_RE.finditer(s.upper())]
        batch_codes = [c for c in all_caps if BATCH_RE.match(c)]
        faculty  = [c for c in all_caps
                    if c not in SKIP_TOKENS
                    and not BATCH_RE.match(c)
                    and (c in FACULTY_MAP or c in self._known_teachers)]
        subject  = clean_subject(s, faculty + batch_codes)
        if not subject or len(subject) < 2 or subject.upper() in SKIP_TOKENS:
            return None
        rooms = _extract_rooms(s)
        return {
            "subject":       subject,
            "type":          "lab" if is_lab else "lec",
            "faculty_codes": faculty,
            "class_div":     "",
            "room":          rooms[0] if rooms else "",
            "raw":           s,
        }

    # ── Internal schedule helpers ─────────────────────────────────────────────

    def _empty_schedule(self) -> dict:
        return {day: {key: [] for key in LECTURE_SLOTS} for day in FIXED_DAYS}

    def _merge_labs(self, schedule: dict) -> dict[str, dict]:
        """
        Detect consecutive lab slots with same subject → record span for response.
        Works on the rich SlotEntry format used by build_*_schedule methods.
        """
        merge_info: dict[str, dict] = {d: {} for d in FIXED_DAYS}
        for day in FIXED_DAYS:
            i = 0
            while i < len(LECTURE_SLOTS):
                key     = LECTURE_SLOTS[i]
                entries = schedule[day].get(key, [])
                if not entries:
                    i += 1
                    continue
                is_lab = any(e["type"] == "lab" for e in entries)
                if is_lab:
                    subj0 = entries[0].get("subject", "")
                    span  = 1
                    while i + span < len(LECTURE_SLOTS):
                        k2   = LECTURE_SLOTS[i + span]
                        ent2 = schedule[day].get(k2, [])
                        if (ent2
                                and any(e["type"] == "lab" for e in ent2)
                                and ent2[0].get("subject", "") == subj0):
                            merge_info[day][k2] = {"skip": True}
                            span += 1
                        else:
                            break
                    merge_info[day][key] = {"span": span}
                i += 1
        return merge_info

    def _schedule_to_response(self, schedule: dict, merge_info: dict) -> dict:
        result = {}
        for day in FIXED_DAYS:
            slots = {}
            for key in LECTURE_SLOTS:
                mi = merge_info.get(day, {}).get(key, {})
                if mi.get("skip"):
                    continue
                entries = schedule[day].get(key, [])
                slots[key] = entries
                if mi.get("span", 1) > 1 and entries:
                    for e in entries:
                        e["span"] = mi["span"]
            result[day] = {"slots": slots}
        return result
